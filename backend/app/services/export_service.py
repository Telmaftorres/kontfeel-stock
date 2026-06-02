import io
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

ZONES = {
    "PALETTE":   "Stockage palette",
    "CHUTE":     "Chute de matière",
    "CONSO":     "Consommables",
    "RACK":      "Racks",
    "RELIQUAT":  "Reliquat",
    "BOX-LOG":   "Box logistique",
    "SHOWROOM":  "Showroom",
    "BE":        "Zone BE",
    "MATERIAU":  "Matériauthèque",
    "FACONNE":   "Façonnage",
    "LOGISTIQUE":"Logistique",
}

ETAT_COLORS = {
    "bon_etat":    "DCFCE7",
    "usage":       "FEF9C3",
    "abime":       "FED7AA",
    "irreparable": "FEE2E2",
}

ARTICLE_HEADERS  = ["Référence", "Nom", "Format", "Unité", "Conditionnement", "Stock actuel", "Stock mini"]
ARTICLE_WIDTHS   = [15, 32, 14, 10, 14, 13, 11]
GOLLECT_HEADERS  = ["Référence", "Nom", "Description", "État", "Stock actuel"]
GOLLECT_WIDTHS   = [15, 32, 38, 14, 13]


def _write_header(ws, headers, widths, bg: str):
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[cell.column_letter].width = w
    ws.row_dimensions[1].height = 26
    ws.freeze_panes = "A2"


def _alternating(r: int) -> str:
    return "EEF2FF" if r % 2 == 0 else "FFFFFF"


MVT_HEADERS = ["Date", "Article", "Nom", "Type", "Quantité", "Stock avant", "Stock après", "Employé", "Étude"]
MVT_WIDTHS  = [18, 16, 30, 10, 10, 12, 12, 16, 16]


def export_all_xlsx(articles: list, gollect_items: list, movements: list | None = None) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    # ── Un onglet par zone ────────────────────────────────────────
    by_zone = defaultdict(list)
    for a in articles:
        by_zone[a.zone].append(a)

    zone_order = [z for z in ZONES if z in by_zone]
    other_zones = [z for z in by_zone if z not in ZONES]

    for zone in zone_order + other_zones:
        label = ZONES.get(zone, zone)[:31]
        ws = wb.create_sheet(title=label)
        ws.sheet_properties.tabColor = "4F46E5"
        _write_header(ws, ARTICLE_HEADERS, ARTICLE_WIDTHS, "4F46E5")

        for r, a in enumerate(by_zone[zone], 2):
            row = [a.ref, a.nom, a.format or "", a.unite, a.conditionnement,
                   float(a.stock_actuel), float(a.stock_mini)]
            fill = PatternFill("solid", fgColor=_alternating(r))
            for c, v in enumerate(row, 1):
                cell = ws.cell(row=r, column=c, value=v)
                cell.fill = fill

        ws.auto_filter.ref = f"A1:G{len(by_zone[zone]) + 1}"

    # ── Onglet Gollect ────────────────────────────────────────────
    ws_g = wb.create_sheet(title="Gollect")
    ws_g.sheet_properties.tabColor = "15803D"
    _write_header(ws_g, GOLLECT_HEADERS, GOLLECT_WIDTHS, "15803D")

    for r, item in enumerate(gollect_items, 2):
        etat = item.etat or ""
        row = [item.ref, item.nom, item.description or "", etat, float(item.stock_actuel)]
        fill = PatternFill("solid", fgColor=ETAT_COLORS.get(etat, _alternating(r)))
        for c, v in enumerate(row, 1):
            cell = ws_g.cell(row=r, column=c, value=v)
            cell.fill = fill

    ws_g.auto_filter.ref = f"A1:E{len(gollect_items) + 1}"

    # ── Onglet Historique mouvements ──────────────────────────────
    if movements:
        ws_m = wb.create_sheet(title="Historique")
        ws_m.sheet_properties.tabColor = "0F172A"
        _write_header(ws_m, MVT_HEADERS, MVT_WIDTHS, "0F172A")

        TYPE_COLORS = {"sortie": "FEE2E2", "entree": "DCFCE7"}
        for r, m in enumerate(movements, 2):
            row = [
                m.created_at.strftime("%d/%m/%Y %H:%M") if m.created_at else "",
                m.article_ref,
                getattr(m, "article_nom", ""),
                m.type,
                float(m.quantite),
                float(m.stock_avant),
                float(m.stock_apres),
                m.employe,
                m.etude_ref or "",
            ]
            fill = PatternFill("solid", fgColor=TYPE_COLORS.get(m.type, _alternating(r)))
            for c, v in enumerate(row, 1):
                ws_m.cell(row=r, column=c, value=v).fill = fill

        ws_m.auto_filter.ref = f"A1:I{len(movements) + 1}"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# Gardés pour compatibilité
def export_articles_xlsx(articles: list) -> bytes:
    return export_all_xlsx(articles, [])


def export_gollect_xlsx(items: list) -> bytes:
    return export_all_xlsx([], items)
